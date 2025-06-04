import models.publisher as publisher_model
import models.book as book_model
import models.supplier as supplier_model
import models.sqlalchemy_config as sqlalchemy_config
import pandas as pd
import os
import warnings
import csv
from datetime import datetime

def recheck_book():
    db = sqlalchemy_config.get_db()

    # publishers = publisher_model.get_all_publishers(db, [], skip=0, limit=10000)
    # publisher_dict = {publisher.publisher_id: publisher for publisher in publishers}
    #
    # suppliers = supplier_model.get_all_suppliers(db, [], skip=0, limit=10000)
    # suppliers_dict = {supplier.supplier_id: supplier for supplier in suppliers}

    file_path = "recheck_book_current_or_need_create/全系統 book_id 校正清單 - 執行_有 isbn 的.csv"

    data = []
    with open(file_path, mode='r', encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            csv_supplier = row["O-供應商"]
            csv_order_number = row["O-單號"]
            csv_order_period = row["O-帳期"]
            csv_book_id = row["O-book_id"]
            csv_isbn = row["O-isbn"]
            csv_title = row["O-書名"]

            result_book = book_model.first_book_by_isbn(db, csv_isbn)
            if (result_book):
                print(f"*** ISBN Matched: {result_book.isbn} == {csv_isbn}")
            else:
                data.append({
                    "O-供應商": csv_supplier,
                    "O-單號": csv_order_number,
                    "O-帳期": csv_order_period,
                    "O-book_id": csv_book_id,
                    "O-isbn": csv_isbn,
                    "O-書名": csv_title,
                })

                print(f"*** ISBN Missing:  {csv_isbn}")
    db.close()

    # 如果有資料，寫入 Excel
    if data:
        file_path = "ISBN Missing.xlsx"
        sheet_name = "比對結果"
        write_to_excel(file_path, sheet_name, data)


def recheck_book_with_isbn():
    db = sqlalchemy_config.get_db()

    publishers = publisher_model.get_all_publishers(db, [], skip=0, limit=10000)
    publisher_dict = {publisher.publisher_id: publisher for publisher in publishers}

    suppliers = supplier_model.get_all_suppliers(db, [], skip=0, limit=10000)
    suppliers_dict = {supplier.supplier_id: supplier for supplier in suppliers}

    file_path = "recheck_book_current_or_need_create/isbn_c1.csv"

    data = []
    with open(file_path, mode='r', encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            csv_book_id = row["book_id"]
            csv_isbn = row["isbn"]
            csv_title = row["書名"]
            csv_publisher = row["出版社"]

            # input(f"*** Processing row: {csv_book_id}, {csv_isbn}, {csv_title}, {csv_publisher}")

            result_book = book_model.get_book_by_isbn(db, csv_isbn)
            if (len(result_book) > 0):
                data.append({
                    "book_id": csv_book_id,
                    "isbn": csv_isbn,
                    "書名": csv_title,
                    "出版社": csv_publisher,
                    "系統數量": len(result_book),
                    "系統 book_id": "",
                    "是否符合": "",
                })
                for book in result_book:
                    check = str(book.book_id) == str(csv_book_id)
                    check = "正確" if check else "異常"

                    data.append({
                        "book_id": "",
                        "isbn": "",
                        "書名": "",
                        "出版社": "",
                        "系統數量": "",
                        "系統 book_id": book.book_id,
                        "是否符合": check,
                    })

                    # print(f"*** ISBN : {book.isbn} == {csv_isbn}, book_id: {book.book_id} == {csv_book_id}")
            else:
                data.append({
                    "book_id": csv_book_id,
                    "isbn": csv_isbn,
                    "書名": csv_title,
                    "出版社": csv_publisher,
                    "系統數量": 0,
                    "系統 book_id": "",
                    "是否符合": "異常",
                })

                # print(f"*** ISBN Missing:  {csv_isbn}")

            if (len(data) > 5000):
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                file_path = f"ISBN_Missing_{timestamp}.xlsx"
                sheet_name = "比對結果"
                print(f"********* Write to Excel {file_path}*********")
                write_to_excel(file_path, sheet_name, data)
                data = []
        # input(f"*** Processing row: {csv_book_id}, {csv_isbn}, {csv_title}, {csv_publisher}")

    db.close()

    # 如果有資料，寫入 Excel
    if data:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        file_path = f"ISBN_Missing_{timestamp}.xlsx"
        sheet_name = "比對結果"
        print(f"********* Write to Excel {file_path}*********")
        write_to_excel(file_path, sheet_name, data)


def recheck_book_with_title():
    db = sqlalchemy_config.get_db()

    # publishers = publisher_model.get_all_publishers(db, [], skip=0, limit=10000)
    # publisher_dict = {publisher.publisher_id: publisher for publisher in publishers}
    #
    # suppliers = supplier_model.get_all_suppliers(db, [], skip=0, limit=10000)
    # suppliers_dict = {supplier.supplier_id: supplier for supplier in suppliers}

    file_path = "recheck_book_current_or_need_create/舊系統書目資料備份_無 isbn - 工作表1.csv"

    data = []
    with open(file_path, mode='r', encoding='utf-8') as file:
        csv_reader = csv.DictReader(file)
        for row in csv_reader:
            csv_book_id = row["book_id"]
            csv_isbn = row["isbn"]
            csv_title = row["書名"]
            csv_publisher = row["出版社"]

            # input(f"*** Processing row: {csv_book_id}, {csv_isbn}, {csv_title}, {csv_publisher}")

            result_book = book_model.get_book_by_title(db, csv_title)
            if (len(result_book) > 0):
                data.append({
                    "book_id": csv_book_id,
                    "isbn": csv_isbn,
                    "書名": csv_title,
                    "出版社": csv_publisher,
                    "系統數量": len(result_book),
                    "系統 book_id": "",
                    "是否符合": "",
                })
                for book in result_book:
                    check = str(book.book_id) == str(csv_book_id)
                    check = "正確" if check else "異常"
                    data.append({
                        "book_id": "",
                        "isbn": "",
                        "書名": "",
                        "出版社": "",
                        "系統數量": "",
                        "系統 book_id": book.book_id,
                        "是否符合": check,
                    })

                    print(f"*** Title : {book.title} == {csv_title}, book_id: {book.book_id} == {csv_book_id}")
            else:
                data.append({
                    "book_id": csv_book_id,
                    "isbn": csv_isbn,
                    "書名": csv_title,
                    "出版社": csv_publisher,
                    "系統數量": 0,
                    "系統 book_id": "",
                    "是否符合": "異常",
                })

                print(f"*** Title Missing:  {csv_title}")

        # input(f"*** Processing row: {csv_book_id}, {csv_isbn}, {csv_title}, {csv_publisher}")

    db.close()

    # 如果有資料，寫入 Excel
    if data:
        file_path = "Title Missing.xlsx"
        sheet_name = "比對結果"
        write_to_excel(file_path, sheet_name, data)


def write_to_excel(file_path, sheet_name, data):
    # 這部分保持不變
    df = pd.DataFrame(data)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        # 如果文件存在，載入 workbook
        book = load_workbook(file_path)

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            # 找到最後一行，將 DataFrame 的內容追加
            for row in dataframe_to_rows(df, index=False, header=False):
                sheet.append(row)
        else:
            # 如果工作表不存在，新增一個表並寫入數據
            sheet = book.create_sheet(sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)

        # 保存文件
        book.save(file_path)

    except FileNotFoundError:
        # 如果文件不存在，直接創建新的 Excel 文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# 禁用警告
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)

if __name__ == '__main__':
    # recheck_book()
    recheck_book_with_isbn()
    # recheck_book_with_title()
