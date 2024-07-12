import models.sqlalchemy_config as sqlalchemy_config
import models.book as book_model
import models.purchase as purchase_model
import models.purchase_item as purchase_item_model
import models.order as order_model
import models.order_item as order_item_model

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import numpy as np
from pprint import pformat

def real_round(num, decimal=0):
    num = float(num)  # Convert to float
    if decimal == 0:
        return int(num + 0.5) if num > 0 else int(num - 0.5)
    else:
        digit_value = 10 ** decimal
        return int(num * digit_value + 0.5) / digit_value if num > 0 else int(num * digit_value - 0.5) / digit_value
def export_orders_to_excel(order, books):
    global row_num

    # for order in orders:
    order_id = order.order_id
    total_price = order.total_price
    current_total_price = 0
    order_row_num = row_num
    sheet[f"A{order_row_num}"] = order_id
    sheet[f"B{order_row_num}"] = total_price
    row_num += 1
    for item in order.order_items:
        order_item_id = item.id
        book = books.get(item.book_id)
        price = book.price
        sale_discount = book.sale_discount
        current_price = real_round(price * sale_discount,0)
        current_total_price += current_price
        # print(
        #     f" Current Price: {current_price}, Book Price: {price}, Book Sale Discount: {book.sale_discount}")

        sheet[f"B{row_num}"] = order_item_id
        sheet[f"C{row_num}"] = item.book_id
        sheet[f"D{row_num}"] = item.price
        sheet[f"E{row_num}"] = item.sale_discount
        sheet[f"F{row_num}"] = current_price
        sheet[f"G{row_num}"] = price
        sheet[f"H{row_num}"] = sale_discount

        # Apply yellow fill if item.price is not equal to current_price
        if item.price != current_price or item.sale_discount != sale_discount:
            for col in range(2, 9):  # Columns B to H
                sheet[f"{get_column_letter(col)}{row_num}"].fill = yellow_fill


            db = sqlalchemy_config.get_db()
            order_item_model.update_order_item_by_id(db, order_item_id,
                                                           {"price":current_price, "sale_discount": sale_discount})

        row_num += 1

    sheet[f"C{order_row_num}"] = current_total_price
    db = sqlalchemy_config.get_db()
    order_model.update_order_by_id(db, order_id,
                                             {"total_price": current_total_price})


def export_purchases_to_excel(purchase, books):
    global row_num

    # for order in orders:
    sheet[f"A{row_num}"] = purchase.purchase_id
    row_num += 1
    for item in purchase.purchase_items:
        book = books.get(item.book_id)
        current_price = book.price
        # print(
        #     f" Current Price: {current_price}, Book Price: {book.price}, Book Sale Discount: {book.sale_discount}")

        purchase_item_id = item.id
        purchase_discount = book.purchase_discount

        sheet[f"B{row_num}"] = item.id
        sheet[f"C{row_num}"] = item.book_id
        sheet[f"D{row_num}"] = item.price
        sheet[f"E{row_num}"] = item.sale_discount
        sheet[f"F{row_num}"] = item.purchase_discount
        sheet[f"G{row_num}"] = current_price
        sheet[f"H{row_num}"] = book.price
        sheet[f"I{row_num}"] = book.sale_discount
        sheet[f"J{row_num}"] = purchase_discount

        # Apply yellow fill if item.price is not equal to current_price
        if item.price != current_price or item.purchase_discount != book.purchase_discount:
            for col in range(2, 11):  # Columns B to J
                sheet[f"{get_column_letter(col)}{row_num}"].fill = yellow_fill

            db = sqlalchemy_config.get_db()
            purchase_item_model.update_purchase_item_by_id(db, purchase_item_id,
                                                           {"price":current_price, "purchase_discount": purchase_discount})


        row_num += 1


def use_get_all_orders():
    db = sqlalchemy_config.get_db()
    page = 1
    page_size = 100
    all_orders = []

    while True:
        paginated_result = order_model.get_paginated_orders(db, [], page=page, page_size=page_size)
        orders = paginated_result.items
        if not orders:
            break

        all_orders.extend(orders)

        for order in orders:
            print(f"Order ID: {order.order_id}")
            book_ids = [item.book_id for item in order.order_items]
            # print(book_ids)
            books = book_model.get_book_by_ids(db, book_ids)
            book_dict = {book.book_id: book for book in books}

            any_printed = False

            for item in order.order_items:
                book = book_dict.get(item.book_id)
                if book:
                    current_price = real_round(book.price * book.sale_discount,0)
                    if item.price != current_price or item.sale_discount != book.sale_discount:
                        print(
                            f"  Order Item ID: {item.book_id}, Book ID: {item.book_id}, "
                            f"Item Price: {item.price}, Current Price: {current_price}, Book Price: {book.price}, "
                            f"Item Sale Discount: {item.sale_discount}, Book Sale Discount: {book.sale_discount}"
                        )
                        any_printed = True
                else:
                    print(f"Book not found: {item.book_id}")
                    input("Press Enter to continue...")
                # print(
                #     f"  Order Item ID: {item.id}, Book ID: {item.book_id}, Price: {item.price}, Sale Discount: {item.sale_discount})")

            if any_printed:
                # for book_id, book in book_dict.items():
                #     print(f"Book ID: {book_id}, Book: {book.price}")

                export_orders_to_excel(order, book_dict)

        page += 1

    total = len(all_orders)
    pages = (total // page_size) + (1 if total % page_size > 0 else 0)

    print(f"Total: {total}, Pages: {pages}")
    print('Done')


def use_get_all_purchases():
    db = sqlalchemy_config.get_db()
    page = 1
    page_size = 100
    all_purchases = []

    while True:
        paginated_result = purchase_model.get_paginated_purchases(db, [], page=page, page_size=page_size)
        purchases = paginated_result.items
        if not purchases:
            break

        all_purchases.extend(purchases)

        for purchase in purchases:
            print(f"Purchase ID: {purchase.purchase_id}")
            book_ids = [item.book_id for item in purchase.purchase_items]
            # print(book_ids)
            books = book_model.get_book_by_ids(db, book_ids)
            book_dict = {book.book_id: book for book in books}

            any_printed = False

            for item in purchase.purchase_items:
                book = book_dict.get(item.book_id)
                if book:
                    current_price = book.price
                    if item.price != current_price or item.purchase_discount != book.purchase_discount:
                        # purchase_item_dict = vars(item)
                        # print(f"*** Purchase Item: {pformat(purchase_item_dict)}")
                        print(
                            f"  Purchase Item ID: {item.id}, Book ID: {item.book_id}, "
                            f"Item Price: {item.price}, Current Price: {current_price}, Book Price: {book.price}, "
                            f"Item Purchase Discount: {item.purchase_discount}, Book Purchase Discount: {book.purchase_discount}"
                        )
                        any_printed = True
                else:
                    print(f"Book not found: {item.book_id}")
                    input("Press Enter to continue...")
                # print(
                #     f"  Purchase Item ID: {item.id}, Book ID: {item.book_id}, Price: {item.price}, Sale Discount: {item.sale_discount})")

            if any_printed:
                # for book_id, book in book_dict.items():
                #     print(f"Book ID: {book_id}, Book: {book.price}")

                export_purchases_to_excel(purchase, book_dict)

        page += 1

    total = len(all_purchases)
    pages = (total // page_size) + (1 if total % page_size > 0 else 0)

    print(f"Total: {total}, Pages: {pages}")
    print('Done')


if __name__ == '__main__':

    # Define row_num as a global variable
    row_num = 2

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Orders"

    # Write headers
    headers = ["訂單編號", "訂單明細編號", "書籍編號", "售出價格（含折扣）", "折扣", "正確售出價格（含折扣）",
               "目前紀錄價錢", "目前紀錄折扣"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Define yellow fill pattern
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # use_get_all_books()
    use_get_all_orders()

    # Save the workbook
    workbook.save("orders.xlsx")

    ###################


    # Define row_num as a global variable
    row_num = 2

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Orders"

    # Write headers
    headers = ["採購單編號", "採購單明細編號", "書籍編號", "採購價格", "售出折扣", "採購折扣", "正確採購價格",
               "目前紀錄價錢", "目前紀錄售出折扣", "目前紀錄採購折扣"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"] = header

    # Define yellow fill pattern
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    use_get_all_purchases()

    # Save the workbook
    workbook.save("purchases.xlsx")
