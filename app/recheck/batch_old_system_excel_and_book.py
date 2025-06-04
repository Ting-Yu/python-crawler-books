import models.publisher as publisher_model
import models.book as book_model
import models.supplier as supplier_model
import models.sqlalchemy_config as sqlalchemy_config
import pandas as pd
import os
import warnings


def safe_read_excel(file_path):
    # 可用的引擎列表，按優先順序排列
    engines = [
        ('openpyxl', {}),
        ('xlrd', {'engine': 'xlrd'}),
        ('xlrd2', {'engine': 'xlrd2'}),
    ]

    for engine, kwargs in engines:
        try:
            # 首先嘗試讀取所有工作表
            excel_data = pd.read_excel(file_path, sheet_name=None, **kwargs)
            return excel_data
        except (ImportError, pd.errors.ParserError, Exception) as e:
            print(f"Failed to read {file_path} with {engine} engine: {str(e)}")
            with open('failed_files.txt', 'a') as f:
                f.write(f"Failed to read {file_path} with {engine} engine: {str(e)}\n")

    # 如果所有引擎都失敗
    print(f"Could not read Excel file: {file_path}")
    with open('failed_files.txt', 'a') as f:
        f.write(f"Could not read Excel file: {file_path}\n")
    return None


def recheck_book():
    db = sqlalchemy_config.get_db()

    publishers = publisher_model.get_all_publishers(db, [], skip=0, limit=10000)
    publisher_dict = {publisher.publisher_id: publisher for publisher in publishers}

    suppliers = supplier_model.get_all_suppliers(db, [], skip=0, limit=10000)
    suppliers_dict = {supplier.supplier_id: supplier for supplier in suppliers}

    folder_path = "batch_old_system_excel"
    all_excels_data = [
        os.path.join(folder_path, f) for f in os.listdir(folder_path)
        if f.endswith(('.xlsx', '.xls', '.xlsb'))
    ]

    file_count = len(all_excels_data)
    print(f"*** All Excels Data: len: {file_count}")
    run_count = 0

    # 用於追蹤處理失敗的檔案
    failed_files = []

    for file_path in all_excels_data:
        print(f"*** Processing Excel File Path: {file_path}")

        # 使用安全的讀取函式
        excel_data = safe_read_excel(file_path)

        if excel_data is None:
            failed_files.append(file_path)
            continue

        data = []

        # 遍歷每個工作表
        for sheet_name, sheet_data in excel_data.items():
            print(f"Processing sheet: {sheet_name}")
            # print(f"Columns: {sheet_data.columns.tolist()}")

            # 檢查必要欄位是否存在
            # required_columns = ['book_id', 'isbn', '供應商', '單號', '帳期', '書名', '進折', '進價', '進量', '小計']
            # missing_columns = [col for col in required_columns if col not in sheet_data.columns]
            #
            # if missing_columns:
            #     input(f"Warning: Missing columns in {sheet_name}: {missing_columns}")
            #     continue

            # 遍歷每一行資料
            for index, row in sheet_data.iterrows():
                print(f"Processing row {index + 1}: {row.to_dict()}")
                try:
                    book_id = row['book_id']
                    isbn = row['isbn']
                    title = row['書名']

                    if book_id is None or pd.isna(book_id):
                        if isbn is None or pd.isna(isbn):
                            if title is None or pd.isna(title):
                                print(f"*** Missing book_id, isbn, and title in row {index}")
                                msg = f"{file_path} - {sheet_name} - Missing book_id, isbn, and title in Processing row {index + 1}: {row.to_dict()}"
                                failed_files.append(msg)
                                result_book = None
                            else:
                                result_book = book_model.first_book_by_title(db, title)
                        else:
                            result_book = book_model.first_book_by_isbn(db, isbn)
                    else:
                        result_book = book_model.get_book_by_id(db, book_id)

                    if result_book and result_book.isbn != isbn:
                        publisher_id = result_book.publisher_id
                        publisher = publisher_dict.get(publisher_id, {})
                        supplier_id = publisher.supplier_id
                        supplier = suppliers_dict.get(supplier_id, {})

                        data.append({
                            "O-供應商": row["供應商"],
                            "O-單號": row["單號"],
                            "O-帳期": row["帳期"],
                            "O-book_id": book_id,
                            "O-isbn": isbn,
                            "O-書名": row["書名"],
                            "O-進折": row["進折"],
                            "O-進價": row["進價"],
                            "O-進量": row["進量"],
                            "O-小計": row["小計"],
                            "<原始資料|系統資料>": "",
                            "N-供應商": supplier.name,
                            "N-book_id": book_id,
                            "N-isbn": result_book.isbn,
                            "N-書名": result_book.title,
                            "N-Sale Discount": result_book.sale_discount,
                            "N-Purchase Discount": result_book.purchase_discount,
                            "N-進價": result_book.price,
                        })
                        print(f"*** ISBN Mismatch: {result_book.isbn} != {isbn}")

                except Exception as e:
                    print(f"Error processing row {index}: {str(e)}")
                    msg = f"{file_path} - {sheet_name} - Missing book_id, isbn, and title in Processing row {index + 1}: {str(e)}"
                    failed_files.append(msg)

        # 如果有資料，寫入 Excel
        if data:
            file_path = "example.xlsx"
            sheet_name = "比對結果"
            write_to_excel(file_path, sheet_name, data)

        run_count += 1
        print(f"*** Run Count: {run_count}/{file_count}")

    db.close()

    # 打印失敗的檔案
    if failed_files:
        print("\n*** Failed to process following files:")
        with open('failed_files.txt', 'w') as f:
            for file in failed_files:
                print(file)
                f.write(file + '\n')


def write_to_excel(file_path, sheet_name, data):
    # 這部分保持不變
    df = pd.DataFrame(data)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        # 如果文件存在，載入 workbook
        book = load_workbook(file_path)

        if sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            # 找到最後一行，將 DataFrame 的內容追加
            for row in dataframe_to_rows(df, index=False, header=False):
                sheet.append(row)
        else:
            # 如果工作表不存在，新增一個表並寫入數據
            sheet = book.create_sheet(sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                sheet.append(row)

        # 保存文件
        book.save(file_path)

    except FileNotFoundError:
        # 如果文件不存在，直接創建新的 Excel 文件
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# 禁用警告
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=RuntimeWarning)

if __name__ == '__main__':
    recheck_book()